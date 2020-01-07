import os
import json
import time
import copy
import PySimpleGUI as sg
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException, ElementClickInterceptedException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from datetime import datetime, timedelta, date

# ----- Declare Global Variables ----- #

# STRUCTURE - data = [ {work_order_number:..., resource:..., excel_row:..., start_date:..., finish_date:...} ... ]
data = []
settings = {
	'excel_file_location': 'Location not set',
	'work_order_submission': True,
	'excel_file_data': {
		'last_m_time': 0
	}
}
cur_entry = 10

# Convert text status into number form
status_number = {
	'In Progress' : 0,
	'' : 0,
	'Filled Out': 1,
	'Approved' : 2,
	'Skipped' : -1,
}

scheduled_task = False

# ----- End of Global Variable Declaration ----- #

# ----- Declare Functions ----- #

def okcanceldialog(title = 'Title', msg='Choose an option', btn_text_one='Ok', btn_text_two='Cancel'):
	lyot = [
		[sg.Text(msg)],
		[sg.Button(button_text=btn_text_one), sg.Button(button_text=btn_text_two)]
	]
	window = sg.Window(title, lyot)
	event, values = window.read()
	window.close()
	return event == btn_text_one

def okdialog(title='Title', msg='Statement', btn_text = 'Ok'):
	lyot = [
		[sg.Text(msg)],
		[sg.Button(button_text=btn_text)]
	]
	window = sg.Window(title, lyot)
	window.close()

def wo_submit_dialog(work_order_index):
	t_header = [
		sg.Text('Work Order Number', size=(20,1), text_color='white', background_color='black'),
		sg.Text('Resource', size=(20,1), text_color='white', background_color='black'),
		sg.Text('Start Time', size=(20,1), text_color='white', background_color='black'),
		sg.Text('Finish Time', size=(20,1), text_color='white', background_color='black')
	]
	l = [
		[sg.Text('Work Order Data Submission', size=(30, 1), justification='center', font=("Helvetica", 14), relief=sg.RELIEF_RIDGE)],
		t_header,
		[
			sg.Text(data[work_order_index]['work_order_number'], size=(20,1)),
			sg.Text(data[work_order_index]['resource'], size=(20,1)),
			sg.Text(data[work_order_index]['start_time'], size=(20,1)),
			sg.Text(data[work_order_index]['finish_time'], size=(20,1))
		],
		[sg.Button('ABORT'), sg.Button('Skip')],
		[sg.Submit()]
	]
	w = sg.Window('Data is about to be submitted', l)
	while True:
		e, v = w.Read()

		if e is None:
			continue

		if e == 'Submit':
			break
		if e == 'ABORT':
			w.close()
			okdialog('Aborting Program', 'Your progress has been saved. Aborting Program.', 'Abort Program')
			driver.close()
			exit()
		if e == 'Skip':
			data[work_order_index]['status'] = 'Skipped'
			if( okcanceldialog('Skip Work Order', 'Are you sure you want to skip the current work order?\nYou can come back to it later.', 'Skip') is True):
				w.close()
				return 1

# Program is getting stuck in infinite loop here, retry can never get to be false
def wait_for_elem(id, wait_duration = 2, err_msg = 'No Error Message Supplied'):
	try:
		WebDriverWait(driver, wait_duration).until(EC.element_to_be_clickable((By.ID, id)))
	except TimeoutException:
		print('Timeout')
		return False
	return True

def click_id(id):
	try:
		driver.find_element_by_id(id).click()
	except NoSuchElementException:
		return False
	except ElementClickInterceptedException:
		return False
	return True

def write_into_elem(id, data):
	driver.find_element_by_id(id).send_keys(str(data))

def wait_for_and_click(id, wait_duration = 10, err_msg = "No Error Message Supplied", counter_limit = 5):
	c = 0
	while not click_id(id):
		if c > counter_limit:
			if okcanceldialog(title="Error", msg="Maximo may be stuck loading, do you want to continue waiting?", btn_text_one='Wait', btn_text_two='Quit'):
				c = 0
			else:
				# Save progress so far
				save_data_to_file(data)
				okdialog(title="Quit", msg="This program will now quit, any progress made up until now has been saved.", btn_text='Quit')
				driver.close()
				exit()
		print('BLOCKED BY WAIT')
		wait_for_elem(id, wait_duration, err_msg)
		time.sleep(0.5)
		c=c+1

def id_exists(id):
	try:
		driver.find_element_by_id(id)
	except NoSuchElementException:
		return False
	return True

def load_settings_file():
	global settings
	# Load settings
	if os.path.exists('mxm_settings.json'):
		f = open('mxm_settings.json')
		temp = load_json(f)
		if not temp:
			okdialog('Settings file not valid', 'Settings file is not valid, reverting to defaults.')
		else:
			settings = temp
	else:
		okdialog('Settings file not found', 'Settings file was not found, reverting to defaults.')

def save_settings():
	global settings
	print("Saving to file : ")
	print(settings)
	save_data_to_file(settings)

def open_settings_dialog(is_sub_menu = False):
	global settings
	print(settings)
	s_layout = [
		[sg.Text('Settings', size=(30, 1), justification='center', font=("Helvetica", 14), relief=sg.RELIEF_RIDGE)],
		[sg.Frame(
			layout= [ 
				[sg.Text('Submission Dialog Popup')],
				[sg.Checkbox('Submission Dialog Popup', size=(20,1), default=settings['work_order_submission'])]
			], title='Options',title_color='red', relief=sg.RELIEF_SUNKEN)],
		[sg.Text('Choose excel file', size=(35, 1))],
		[sg.Text('Excel Directory', size=(15, 1), auto_size_text=False, justification='right'),
			sg.InputText(settings['excel_file_location']), sg.FileBrowse()],
		[sg.Button('Apply')]
	]

	s_win = sg.Window('Settings Window', s_layout)

	while True:
		s_ev, s_val = s_win.Read()

		if s_ev is None:
			if is_sub_menu:
				return True
			exit()
		if s_ev == "Apply":
			settings['work_order_submission'] = s_val[0]
			settings['excel_file_location'] = s_val[1]
			save_data_to_file(settings, 'mxm_settings.json')
			s_win.close()
			if is_sub_menu:
				return True
			break

def load_excel_file():
	global data
	for cur_row in range(cur_entry, 71):

		won = ws.cell(column=4, row=cur_row).value
		resource = ""

		# Check if this row has values assigned
		if(won != None):
			# Find resource allocated to Work Order
			for d in range(6):
				temp_resource = ws.cell(column=(6+(2*d)), row=cur_row).value
				if temp_resource:
					resource = temp_resource.replace(' ', '')
					print(resource)
					break
				else:
					# No resource found in this cell
					continue
				
			temp = {
				'work_order_number': won,
				'resource': resource,
				'excel_row': cur_row,
				'start_time': next_monday.strftime("%m/%d/%Y %I:%M %p"),
				'finish_time': next_sunday.strftime("%m/%d/%Y %I:%M %p"),
				'status': ''
			}
			data.append(temp)
	save_data_to_tracker()

def load_entries_file():
	global data
	# Load data from entries file
	if os.path.exists('mxm_data_tracker.json'):
		f = open('mxm_data_tracker.json')
		temp = load_json(f)
		if not temp:
			okdialog('Data Tracker file not valid', 'Data Tracker file is not valid.')
			return False
		else:
			data = temp[1:]
			print(temp)
			print(data)
	else:
		okdialog('Data Tracker file not found', 'Data Tracker file was not found.')
		return False
	return True

def load_json(to_be_loaded):
	try:
		return json.load(to_be_loaded)
	except ValueError:
		return False

def save_data_to_file(d, filename = 'mxm_settings.json'):
	f = open(filename, 'w+')
	json.dump(d, f)

def save_data_to_tracker(filename = 'mxm_data_tracker.json'):
	global data
	temp = []
	temp = copy.deepcopy(data)
	temp.insert(0,settings['excel_file_data']['last_m_time'])
	save_data_to_file(temp, filename)

def navigate_to_work_order(work_order_index):
	global data
	# Navigate to the home page
	wait_for_and_click('titlebar-tb_homeButton')
	time.sleep(1)
	
	# Navigate from home page to work order search
	wait_for_and_click('m7f8f3e49_ns_menu_WO_MODULE_a_tnode')
	time.sleep(0.25)
	wait_for_and_click('m7f8f3e49_ns_menu_WO_MODULE_sub_changeapp_WOTRACK')

	# Enter work order into search field
	print('Entering work order into field')
	wait_for_and_click('m6a7dfd2f_tfrow_[C:1]_txt-tb', err_msg="Could not find work order")
	write_into_elem('m6a7dfd2f_tfrow_[C:1]_txt-tb', data[work_order_index]['work_order_number'])

	# Click search icon
	print('Searching for work order')
	wait_for_and_click('m6a7dfd2f-img5')

	# Wait until Work Order loads
	if not wait_for_elem('m6a7dfd2f_tdrow_[C:1]_ttxt-lb[R:0]'):
		# work Order was not listed
		okdialog('Work order could not be found')
		data[work_order_index]['status'] = 'Skipped'
		return True

	time.sleep(1)

	# Open Word Order edit page
	print('Opening work order edit window')
	wait_for_and_click('m6a7dfd2f_tdrow_[C:1]_ttxt-lb[R:0]')

def generate_report():
	# ----------------------- #
	# --- Generate Report --- #
	# ----------------------- #

	# Navigate to the home page
	wait_for_and_click('titlebar-tb_homeButton')
	time.sleep(1)

	# Navigate from home page to work order search
	wait_for_and_click('m7f8f3e49_ns_menu_WO_MODULE_a_tnode')
	time.sleep(0.25)
	wait_for_and_click('m7f8f3e49_ns_menu_WO_MODULE_sub_changeapp_WOTRACK')

	# Open Advanced Search
	wait_for_and_click('m68d8715f-tbb_text')
	time.sleep(1)

	# Enter "Work Type"
	wait_for_and_click('med325893-tb')
	write_into_elem('med325893-tb', '=PM')

	# Enter "Work Group"
	wait_for_and_click('mafbecfc3-tb')
	write_into_elem('mafbecfc3-tb', '=MC')

	# Enter 'Target Start' date and time
	wait_for_and_click('m3cdc438b-tb')
	write_into_elem('m3cdc438b-tb', (next_monday - timedelta(days=14)).strftime("%m/%d/%Y %I:%M %p"))

	# Enter 'Target Finish' date and time
	wait_for_and_click('md8d7fe4c-tb')
	write_into_elem('md8d7fe4c-tb', (next_sunday - timedelta(days=14)).strftime("%m/%d/%Y %I:%M %p"))

	# Find results
	wait_for_and_click('m4fd840b0-pb')

	# Run Reports
	wait_for_and_click('m74daaf83_ns_menu_RUNREPORTS_OPTION_a_tnode')

	# Click 'Create Report' Button
	wait_for_and_click('me073773f-pb')

	# Enter report title name
	wait_for_and_click('m287e5525-tb')
	write_into_elem('m287e5525-tb',  date.today().strftime('%d-%b-%Y %I:%M - Generated Report') )

	# Navigate to 'Content' tab
	wait_for_and_click('m90576a63-tab_anchor')
	time.sleep(1)

	rpt_search_words = ['lead', 'scheduled finish']

	for search_word in rpt_search_words:
		# Look for additional fields to add to report
		wait_for_and_click('m83ceef56_tfrow_[C:1]_txt-tb')
		write_into_elem('m83ceef56_tfrow_[C:1]_txt-tb', search_word)

		# Search for enterd word
		wait_for_and_click('m83ceef56-ti2_img')

		# Select required additional report field
		wait_for_and_click('m83ceef56_tdrow_[C:1]_ttxt-lb[R:0]')
		time.sleep(1)

	# Click 'Run and Save Completed Report'
	wait_for_and_click('m8dcd759e-pb')

load_settings_file()
open_settings_dialog()

# ----- Load workbook ----- #
wb = load_workbook(settings['excel_file_location'])
ws = wb.active

# Get next mondays date
today = date.today()
next_monday = datetime(today.year, today.month, today.day, 0, 0, 0)
while next_monday.weekday() != 0:
	next_monday += timedelta(1)   
# Get the following sundays date
next_sunday = next_monday + timedelta(6)

next_monday += timedelta(hours=5)
next_sunday += timedelta(hours=23.75)

# Load workbook last modified time
lastmtime = os.path.getmtime(settings['excel_file_location'])
if settings['excel_file_data']['last_m_time'] != lastmtime:
	mdfy_layout = [
		[sg.Text('Your excel file seems to have been modified since this program was last ran.')],
		[sg.Button('Reload Data'), sg.Button('Keep old data')]
	]
	mdfy_win = sg.Window('Reload Data?', mdfy_layout)
	while True:
		mdf_e, mdf_val = mdfy_win.Read()
		if mdf_e is None:
			continue
		if mdf_e == 'Reload Data':
			print('Reloading new data')

			# Update settings last modified time for excel file
			settings['excel_file_data']['last_m_time'] = lastmtime

			# Load excel data into program
			load_excel_file()

			# Save current settings
			save_settings()
			break
		if mdf_e == 'Keep old data':
			print('keeping old data')
			if load_entries_file():
				print('Entries file has been sucessfully loaded')
			else:
				# TODO: Handle when entries tracker file is not loaded sucessfully
				print('Entries file was not loaded')
				okdialog('Data Tracker File not loaded', 'The data tracker file was not able to be loaded.')
				# TODO: Add dialog that will allow the user to delete the data_tracker.json file if it has gotten corrupted
			break
	mdfy_win.close()
	# Ask user if they want to regenerate their excel data file
	# In case there were edits made the program will update incorrect values becuse the work order layout could be different
	print('Last modified time is the same as settings variable says')
else:
	load_entries_file()

# Define layout for base startup menu
layout = [
	[sg.Text('Work Order Automator', size=(30, 1), justification='center', font=("Helvetica", 14), relief=sg.RELIEF_RIDGE)],
	[sg.Text('Work Order Count'), sg.Text(str(len(data))), sg.Button('View Work Order data')],
	[sg.Button('Settings')],
	[sg.Button('Generate Report')],
	[sg.Button('Start'), sg.Exit()]
]

win = sg.Window('Window 1', layout)

while True:
	ev, vals = win.Read()

	if ev is None or ev == "Exit":
		exit()

	if ev == "Start":
		break
	if ev == "View Work Order data":
		table_header = [
			sg.Text('Status', size=(20,1), text_color='white', background_color='black'),
			sg.Text('Work Order Number', size=(20,1), text_color='white', background_color='black'),
			sg.Text('Resource', size=(20,1), text_color='white', background_color='black'),
			sg.Text('Start Time', size=(20,1), text_color='white', background_color='black'),
			sg.Text('Finish Time', size=(20,1), text_color='white', background_color='black')
		]
		layout2 = [
			[sg.Text('Work Order Listings', size=(30, 1), justification='center', font=("Helvetica", 14), relief=sg.RELIEF_RIDGE)],
			table_header
		]
		
		for ind in range(len(data)):
			row = []
			if data[ind]['status'] == 'Approved':
				row.append(sg.Text('[Approved   ]', size=(20,1)))
			elif data[ind]['status'] == 'In Progress':
				row.append(sg.Text('[In Progress]', size=(20,1)))
			elif data[ind]['status'] == 'Filled Out':
				row.append(sg.Text('[Filled Out ]', size=(20,1)))
			elif data[ind]['status'] == 'Skipped':
				row.append(sg.Text('[Skipped    ]', size=(20,1)))
			else:
				row.append(sg.Text('[           ]', size=(20,1)))
			row.extend( [sg.Text(data[ind]['work_order_number'], size=(20,1)), sg.Text(data[ind]['resource'], size=(20,1)), sg.Text(data[ind]['start_time'], size=(20,1)), sg.Text(data[ind]['finish_time'], size=(20,1))] )
			layout2.append( row )
		layout2.append( [sg.Ok()] )

		win.hide()
		win2 = sg.Window("Work Order Data", layout2)

		ev2, vals2 = win2.Read()

		win2.close()
		win.UnHide()
	if ev == 'Settings':
		open_settings_dialog(True)
	if ev == 'Generate Report':
		scheduled_task = 'gr'
		break

win.hide()

# Setup Web Driver
driver = webdriver.Chrome()
#driver = webdriver.Ie()
# Open Maximo in Chrome
driver.get('https://ecmms.jnj.com/maximo/webclient/login/login.jsp?appservauth=true')

# Prompt user to log in
while True:
	answer = okcanceldialog("Question", "Press \"OK\" after you have logged in and Maximo home page is displayed")
	print(answer)
	if answer is True:
		if id_exists('m7f8f3e49_ns_menu_WO_MODULE_a_tnode'):
			break
	else:
		driver.close()
		exit()

if scheduled_task == 'gr':
	generate_report()
	okdialog('Report Generated', 'Your report has been generated!')
	exit()

# Fill data for each work order
for work_order_index in range(len(data)):

	# Structure data to be dumped into file
	save_data_to_tracker()

	status = status_number[data[work_order_index]['status']]
	starting_status = True

	# -------------------- #
	# ------ Stage 1 ----- #
	# -------------------- #

	if( status < 1 ):

		if navigate_to_work_order(work_order_index):
			continue

		# Enter scheduled start date
		print('Entering start date')
		print(str(next_monday))
		wait_for_and_click('m8b12679c-tb')
		write_into_elem('m8b12679c-tb', data[work_order_index]['start_time'])

		# Enter scheduled finish date
		print('entering finish date')
		print(str(next_sunday))
		wait_for_and_click('m1576f23f-tb')
		write_into_elem('m1576f23f-tb', data[work_order_index]['finish_time'])

		# Enter lead resource
		print('Entering lead resource')
		wait_for_and_click('mec969533-img')

		# Open leader search pane
		wait_for_and_click('NORMAL_normal0_a')

		wait_for_and_click('lookup_page1_tfrow_[C:0]_txt-tb', err_msg="Leader could not be found.")

		# Enter leader name
		wait_for_elem('lookup_page1_tfrow_[C:0]_txt-tb')
		write_into_elem('lookup_page1_tfrow_[C:0]_txt-tb', data[work_order_index]['resource'])

		# Search for leader
		wait_for_and_click('lookup_page1-ti2_img')

		# Wait for leader list to load
		wait_for_and_click('lookup_page1_tdrow_[C:0]-c[R:0]')

		time.sleep(3)

		# Check if the user wants to have a dialog box appear to ask them to verify before anything is saved
		if settings['work_order_submission']:
			# Display submission verification dialog box
			if( wo_submit_dialog(work_order_index) == 1):
				# Set current work order status to Skipped
				data[work_order_index]['status'] = 'Skipped'
				continue

		# Save Work Order
		print('Saving Work Order')
		if( okcanceldialog(title="Question", msg="Do you want to save the changes that have been made?") ):
			wait_for_and_click('toolactions_SAVE-tbb_image')
			time.sleep(3)
		else:
			# Skip or exit program
			exit()

		# Update status
		data[work_order_index]['status'] = 'Filled Out'
		status = status+1
		starting_status = False

	# --------------------- #
	# ------ Stage 2 ------ #
	# --------------------- #

	if (status < 2):

		if starting_status:
			starting_status = False
			if navigate_to_work_order(work_order_index):
				continue

		# Change status of the WorkOrder
		# Will automatically save the work order before opening this link
		wait_for_and_click('md86fe08f_ns_menu_STATUS_OPTION_a_tnode')

		# Click the drop down for work order status
		wait_for_and_click('mc927149a-img')

		# Check if 'Approved' option is available
		if( not wait_for_elem('menu0_APPR_OPTION_a') ):
			# 'Approved' option is not available
			data['work_order_index'] = 'Skipped'
			okdialog('No approve option', 'Could not change status of Work Order to Approved because \'Approved\' was not an option.')
			continue

		# Select 'Approved' option drop-down from menu
		wait_for_and_click('menu0_APPR_OPTION_a')

		# Select 'OK' to lock in new status change
		wait_for_and_click('m60bd6d91-pb')

		# Page will automatically saved after a few seconds after pressing previous 'OK' button
		data[work_order_index]['status'] = 'Approved'

		# Navigate to work order home page
		wait_for_and_click('m397b0593-tabs_middle')

		# Update work order status
		data[work_order_index]['status'] = 'Approved'

	# Save progress to data_tracker.json
	save_data_to_tracker()

driver.close()
okdialog('Complete', 'Finished Operation. \nPlease ensure to complete any skipped work orders.')