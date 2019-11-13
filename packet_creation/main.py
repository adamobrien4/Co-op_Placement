import time
import os
import re

import pyautogui as pag
from openpyxl import load_workbook
from shutil import copyfile

asset = ""
mat_num = ""
plm_num = ""
op_code = ""
machine_program = ""
revision = ""
packet_name = ""

wb = load_workbook('fanuc.xlsx')
ws = wb.active

# pag.moveTo(370, 15)
# pag.click()

def move_click_type(x, y, string):
    pag.moveTo(x, y, duration=0.1)
    pag.click()
    time.sleep(0.1)
    pag.typewrite(str(string))

def fill_packet_details(x):    
    move_click_type(205, 250, packet_name)
    move_click_type(940, 250, op_code)
    move_click_type(200, 280, mat_num)

    # Enter Workplace
    pag.moveTo(744, 277, duration=0.2)
    pag.click()

    time.sleep(1)
    
    pag.moveTo(974, 590, duration=0.2)
    pag.click()
    time.sleep(0.1)
    pag.typewrite(asset)
    time.sleep(0.1)
    pag.press("enter")

    time.sleep(0.25)

    pag.moveTo(1050, 700, duration=0.2)
    pag.click()

    time.sleep(1)

    # Revision
    move_click_type(630, 312, revision)
    # PLM number
    move_click_type(975, 312, plm_num)

    save_changes()
    time.sleep(1)

def save_changes():
    # Save
    pag.moveTo(50, 165, duration=0.1)
    pag.click()

for x in range(189, 200):

    asset = ws.cell(row=x, column=2).value
    mat_num = ws.cell(row=x, column=3).value
    plm_num = ws.cell(row=x, column=8).value
    op_code = ws.cell(row=x, column=10).value
    machine_program = str(ws.cell(row=x, column=11).value)
    revision = ws.cell(row=x, column=9).value
    packet_name = asset+"_"+str(op_code)+"_"+str(mat_num)

    dirroot = "L:\Common\Personal\Ash Chilakwad\File Compare - 11062019/" + asset + "/temp/"
    
    # Search for packets
    pag.moveTo(120, 170, duration=0.2)
    pag.click()

    time.sleep(1.75)

    # Create new packet
    pag.moveTo(650, 240, duration=0.2)
    pag.rightClick()
    pag.moveTo(660, 250, duration=0.2)
    pag.click()

    # Wait for new packet page to open
    time.sleep(2)
        
    fill_packet_details(x)

    # Find current enteries file(s)
    mp_split = machine_program.split(",")
    print(len(mp_split))

    machine_file_dir = []

    if( len(mp_split) == 1 ):
        file_noext = re.sub(r'\W+', '', mp_split[0].split(".")[0])
        source = dirroot + re.sub(r'\W+', '', mp_split[0]) + ".tmp"
        dest = dirroot + file_noext + ".nc"
        copyfile(source, dest)
        machine_file_dir.append(file_noext + ".nc")
    elif( len(mp_split) == 2 ):
        for i in range(len(mp_split)):
            file_noext = re.sub(r'\W+', '', mp_split[i].split(".")[0])
            source = dirroot + re.sub(r'\W+', '', mp_split[i]) + ".tmp"
            dest = dirroot + file_noext + ".nc" + str(i+1)
            print(source)
            print(dest)
            copyfile(source, dest)
            if(i == 0):
                machine_file_dir.append(file_noext + ".nc1")
            elif(i == 1):
                machine_file_dir.append(file_noext + ".nc2")
    else:
        print("Machine files greater than 2 what?")
        exit()

    for i in range(len(machine_file_dir)):
        # Open upload dialog window
        pag.moveTo(420, 700, duration=0.2)
        pag.rightClick()
        pag.moveTo(430, 710, duration=0.2)
        pag.click()

        time.sleep(1)

        # Select file(s)
        pag.moveTo(1180, 500, duration=0.1)
        pag.click()

        time.sleep(3.25)

        # Enter correct address in address bar
        if(i == 0):
            pag.moveTo(120, 50, duration=0.2)
            pag.click()
            time.sleep(0.1)
            pag.typewrite(dirroot)
            pag.press("enter")

        # Enter file name
        pag.moveTo(300, 475, duration=0.1)
        pag.click()
        time.sleep(0.1)
        pag.typewrite(machine_file_dir[i])

        pag.moveTo(790, 510, duration=0.1)
        pag.click()

        time.sleep(3)

        pag.moveTo(1130, 690, duration=0.25)
        pag.click()

        time.sleep(1)

        # Enter file revision and plm number
        '''pag.moveTo(750, 510 + (i*30))
        pag.click()
        time.sleep(0.2)
        pag.typewrite(str(revision))

        move_click_type(850, 510 + (i*30), plm_num)'''
        
    save_changes()
    time.sleep(1.25)

    pag.moveTo(440, 125, 0.2)
    pag.click()

    time.sleep(2)
